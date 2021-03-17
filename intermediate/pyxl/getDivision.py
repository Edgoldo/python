import collections
import unidecode

from openpyxl import Workbook, load_workbook
## Carga del archivo División Política de Panamá
division = load_workbook(filename = 'DIVISION POLITICA DE PANAMA.xlsx')
## Carga del archivo MasterEstado
estados = load_workbook(filename = 'MasterEstado.xlsx')
sheetDiv = division['Hoja1']
sheetEst = estados['MasterEstado']
## Obtención de las columnas del archivo División
divA = list(sheetDiv['A'])
divB = list(sheetDiv['B'])
divC = list(sheetDiv['C'])
## Obtención de la columna del archivo MasterEstado
estB = list(sheetEst['B'])
#print(len(divA), len(divB), len(divC))

states = {}
unaccentedStates = []
for cell in divA:
	if cell.value != None and cell.value != "PROVINCIAS Y COMARCAS" and cell.value != "                                                                                                                           DIVISIÓN POLÍTICA DE PANAMÁ":
		states[divA.index(cell)] = cell.value
		unaccentedStates.append(unidecode.unidecode(cell.value))
val = states[673]
del states[673]
states[674] = val
val = states[682]
del states[682]
states[683] = val
val = states[689]
del states[689]
states[690] = val
states = collections.OrderedDict(sorted(states.items()))

defStates = {}
for cell in estB:
	if cell.value in unaccentedStates or cell.value in states.values():
		defStates[estB.index(cell)] = cell.value

index = 4001
for state in states.values():
	unaccentedState = unidecode.unidecode(state)
	if state not in defStates.values() and unaccentedState not in defStates.values():
		defStates[index] = state
		index += 1
del defStates[1719]
defStates = collections.OrderedDict(sorted(defStates.items()))

cities = {}
for cell in divB:
	if cell.value != None and cell.value != "Distrito" and cell.value != "Comarcas" and cell.value != "(Ninguno)":
		cities[divB.index(cell)] = cell.value
del cities[446]
cities =  collections.OrderedDict(sorted(cities.items()))

defCities = {11: "PANAMÁ"}
index = 208
for city in cities.values():
	defCities[index] = city
	index += 1
defCities =  collections.OrderedDict(sorted(defCities.items()))

sectors = {}
for cell in divC:
	if cell.value != None and cell.value != "Corregimientos":
		sectors[divC.index(cell)] = cell.value
sectors =  collections.OrderedDict(sorted(sectors.items()))

defSectors = {11: "PANAMÁ"}
index = 1879
for sector in sectors.values():
	defSectors[index] = sector
	index += 1
defSectors =  collections.OrderedDict(sorted(defSectors.items()))

print("Estados ", len(states), ": ", states)

#print("Estados no acentuados ", len(unaccentedStates), ": ", unaccentedStates)

print("Estados definitivos ", len(defStates), ": ", defStates)

#print("Ciudades ", len(cities), ": ", cities)

#print("Ciudades definitivas ", len(defCities), ": ", defCities)

#print("Municipios ", len(defSectors), ": ", defSectors)

"""
Creación del archivo MasterEstado
wbStates = Workbook()
wsStates = wbStates.create_sheet("MasterEstado", 0)
#wsStates.title = "MasterEstado.xlsx"

headTable = ["idMasterEstado", "nombre", "status", "ultimoUsuario", "ultimaFecha", "MasterPais", "MasterMiscelaneoEstado"]
i = 1
for val in headTable:
	row = wsStates.cell(row=i, column=headTable.index(val)+1, value=val)
i = 2
for key, state in defStates.items():
	if key >= 4001:
		headTable = [key, state.upper(), 1, "anon.", "2017-06-29 09:42:02", 174, 12]
		for val in headTable:
			row = wsStates.cell(row=i, column=headTable.index(val)+1, value=val)
		i += 1
wbStates.save("definitivos/MasterEstado.xlsx")
"""

"""
Creación del archivo MasterCiudad
wbCities = Workbook()
wsCities = wbCities.create_sheet("MasterCiudad", 0)
#wsStates.title = "MasterEstado.xlsx"

headTable = ["idMasterCiudad", "nombre", "status", "ultimoUsuario", "ultimaFecha", "MasterEstado", "MasterMiscelaneoEstado"]
i = 1
for val in headTable:
	row = wsCities.cell(row=i, column=headTable.index(val)+1, value=val)

i = 2
index = 208
for key, city in cities.items():
	try:
		if states[key]:
			state = states[key]
		for k, s in defStates.items():
			if (s == state or s == unidecode.unidecode(state)):
				state = k
	except:
		pass
	headTable = [index, city, 1, "daniel.munoz@solutecsystems.com", "0000-00-00 00:00:00", state, 12]
	for val in headTable:
		row = wsCities.cell(row=i, column=headTable.index(val)+1, value=val)
	index += 1
	i += 1

wbCities.save("definitivos/MasterCiudad.xlsx")
"""

wbSectors = Workbook()
wsSectors = wbSectors.create_sheet("MasterMunicipio", 0)
#wsStates.title = "MasterEstado.xlsx"

headTable = ["idMasterMunicipio", "nombre", "codPostal", "status", "ultimoUsuario", "ultimaFecha", "MasterCiudad", "MasterMiscelaneoEstado"]
i = 1
for val in headTable:
	row = wsSectors.cell(row=i, column=headTable.index(val)+1, value=val)

i = 2
index = 1879
for key, sector in sectors.items():
	try:
		if cities[key]:
			city = cities[key]
		for k, c in defCities.items():
			if (c == city):
				city = k
	except:
		pass
	headTable = [index, sector, 0, 1, "daniel.munoz@solutecsystems.com", "0000-00-00 00:00:00", city, 12]
	for val in headTable:
		row = wsSectors.cell(row=i, column=headTable.index(val)+1, value=val)
	index += 1
	i += 1

wbSectors.save("definitivos/MasterMunicipio.xlsx")